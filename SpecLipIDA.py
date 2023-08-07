import openpyxl

# Open workbook
wb1 = openpyxl.load_workbook('library.xlsx')
wb2 = openpyxl.load_workbook('data list.xlsx')

# Get worksheet
ws1 = wb1.active
ws2 = wb2.active

# Create a new worksheet
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Iterate through each row of ws1
for row1 in ws1.iter_rows(min_row=1, max_row=ws1.max_row):

    # Matching flag bit
    match_flag = False

    # Iterate through each row of ws2
    for row2 in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
        # Judge whether the cell value is None
        if row1[0].value is None or row2[0].value is None:
            continue

        # Comparison of MS1
        if abs(row1[0].value - row2[0].value) < 0.02:

            # Compare MS1 and MS2 of ws1 in turn
            for i in range(1, len(row1)-1):

                # Compare in turn the MS2 corresponding to MS1 in ws2
                for j in range(1, len(row2)):
                    # Judge whether the cell value is None
                    if row1[i].value is None or row2[j].value is None:
                        continue
                    # Judge if it is a numeric type
                    elif isinstance(row1[i].value, (int, float)) and isinstance(row2[j].value, (int, float)):
                        # Comparison of MS2
                        if abs(row1[i].value - row2[j].value) < 0.02:
                            break  # Match successful, jump out of inner loop

                else:
                    # If the match fails, modify the match flag bit and continue to the next line of the comparison
                    match_flag = False
                    break  # Jump out of outer loop

            else:
                # If both MS1 and MS2 are matched successfully, write the compound name of the corresponding row in ws1 to the new worksheet and modify the match flag bit
                new_ws.append([row1[len(row1)-1].value])
                match_flag = True
                break  # Successful match, jump out of outer loop

    # If no match is successful, the null value is written to the new worksheet
    # if not match_flag:
    #     new_ws.append([""])

# Save workbook
new_wb.save('identification result.xlsx')
